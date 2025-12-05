from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# Sheet 1: Competitor Overview
sheet = wb.active
sheet.title = "Competitor Overview"

# Headers
headers = [
    "Company", "Category", "Founded", "HQ Location", "Funding Stage", 
    "Total Funding", "Last Round", "Last Round Date", "Valuation", 
    "Employees", "Key Products", "Deployment", "Primary Use Cases",
    "Differentiators", "Website", "Last Updated"
]

sheet.append(headers)

# Header formatting
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Competitor data
competitors = [
    ["Nightfall AI", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.nightfallai.com", datetime.now().strftime('%Y-%m-%d')],
    ["Varonis", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.varonis.com", datetime.now().strftime('%Y-%m-%d')],
    ["Forcepoint", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.forcepoint.com", datetime.now().strftime('%Y-%m-%d')],
    ["Digital Guardian", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.digitalguardian.com", datetime.now().strftime('%Y-%m-%d')],
    ["Proofpoint", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.proofpoint.com", datetime.now().strftime('%Y-%m-%d')],
    ["Symantec DLP", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.symantecdlp.com", datetime.now().strftime('%Y-%m-%d')],
    ["Cyera", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.cyera.com", datetime.now().strftime('%Y-%m-%d')],
    ["BigID", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.bigid.com", datetime.now().strftime('%Y-%m-%d')],
    ["Sentra", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.sentra.com", datetime.now().strftime('%Y-%m-%d')],
    ["Normalyze", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.normalyze.com", datetime.now().strftime('%Y-%m-%d')],
    ["Laminar", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.laminar.com", datetime.now().strftime('%Y-%m-%d')],
    ["Dig Security", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.digsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Polar Security", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.polarsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Securiti.ai", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.securiti.ai.com", datetime.now().strftime('%Y-%m-%d')],
    ["Netskope", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.netskope.com", datetime.now().strftime('%Y-%m-%d')],
    ["Zscaler", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.zscaler.com", datetime.now().strftime('%Y-%m-%d')],
    ["Skyhigh Security", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.skyhighsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Wiz", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.wiz.com", datetime.now().strftime('%Y-%m-%d')],
    ["Lacework", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.lacework.com", datetime.now().strftime('%Y-%m-%d')],
    ["Orca Security", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.orcasecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["CrowdStrike", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.crowdstrike.com", datetime.now().strftime('%Y-%m-%d')],
    ["SentinelOne", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.sentinelone.com", datetime.now().strftime('%Y-%m-%d')],
    ["Code42", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.code42.com", datetime.now().strftime('%Y-%m-%d')],
    ["SquareX", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.squarex.com", datetime.now().strftime('%Y-%m-%d')],
    ["Island", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.island.com", datetime.now().strftime('%Y-%m-%d')],
    ["Talon", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.talon.com", datetime.now().strftime('%Y-%m-%d')],
    ["LayerX Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.layerxsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Seraphic Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.seraphicsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Surf Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.surfsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Menlo Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.menlosecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Camunda", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.camunda.com", datetime.now().strftime('%Y-%m-%d')],
    ["ProcessMaker", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.processmaker.com", datetime.now().strftime('%Y-%m-%d')],
    ["Appian", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.appian.com", datetime.now().strftime('%Y-%m-%d')],
    ["Pega", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.pega.com", datetime.now().strftime('%Y-%m-%d')],
    ["Nintex", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.nintex.com", datetime.now().strftime('%Y-%m-%d')],
    ["Bizagi", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.bizagi.com", datetime.now().strftime('%Y-%m-%d')],
    ["Flowable", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.flowable.com", datetime.now().strftime('%Y-%m-%d')],
    ["Workato", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.workato.com", datetime.now().strftime('%Y-%m-%d')],
    ["Lakera", "AI Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.lakera.com", datetime.now().strftime('%Y-%m-%d')],
    ["Veza", "Other", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.veza.com", datetime.now().strftime('%Y-%m-%d')],
    ["Microsoft Purview", "Other", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.microsoftpurview.com", datetime.now().strftime('%Y-%m-%d')],
],
    ["Varonis", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.varonis.com", datetime.now().strftime('%Y-%m-%d')],
    ["Forcepoint", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.forcepoint.com", datetime.now().strftime('%Y-%m-%d')],
    ["Digital Guardian", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.digitalguardian.com", datetime.now().strftime('%Y-%m-%d')],
    ["Proofpoint", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.proofpoint.com", datetime.now().strftime('%Y-%m-%d')],
    ["Symantec DLP", "DLP", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.symantecdlp.com", datetime.now().strftime('%Y-%m-%d')],
    ["Cyera", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.cyera.com", datetime.now().strftime('%Y-%m-%d')],
    ["BigID", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.bigid.com", datetime.now().strftime('%Y-%m-%d')],
    ["Sentra", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.sentra.com", datetime.now().strftime('%Y-%m-%d')],
    ["Normalyze", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.normalyze.com", datetime.now().strftime('%Y-%m-%d')],
    ["Laminar", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.laminar.com", datetime.now().strftime('%Y-%m-%d')],
    ["Dig Security", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.digsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Polar Security", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.polarsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Securiti.ai", "DSPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.securiti.ai.com", datetime.now().strftime('%Y-%m-%d')],
    ["Netskope", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.netskope.com", datetime.now().strftime('%Y-%m-%d')],
    ["Zscaler", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.zscaler.com", datetime.now().strftime('%Y-%m-%d')],
    ["Skyhigh Security", "CASB/SASE", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.skyhighsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Wiz", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.wiz.com", datetime.now().strftime('%Y-%m-%d')],
    ["Lacework", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.lacework.com", datetime.now().strftime('%Y-%m-%d')],
    ["Orca Security", "Cloud Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.orcasecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["CrowdStrike", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.crowdstrike.com", datetime.now().strftime('%Y-%m-%d')],
    ["SentinelOne", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.sentinelone.com", datetime.now().strftime('%Y-%m-%d')],
    ["Code42", "Endpoint/EDR", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.code42.com", datetime.now().strftime('%Y-%m-%d')],
    ["SquareX", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.squarex.com", datetime.now().strftime('%Y-%m-%d')],
    ["Island", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.island.com", datetime.now().strftime('%Y-%m-%d')],
    ["Talon", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.talon.com", datetime.now().strftime('%Y-%m-%d')],
    ["LayerX Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.layerxsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Seraphic Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.seraphicsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Surf Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.surfsecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Menlo Security", "Browser Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.menlosecurity.com", datetime.now().strftime('%Y-%m-%d')],
    ["Camunda", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.camunda.com", datetime.now().strftime('%Y-%m-%d')],
    ["ProcessMaker", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.processmaker.com", datetime.now().strftime('%Y-%m-%d')],
    ["Appian", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.appian.com", datetime.now().strftime('%Y-%m-%d')],
    ["Pega", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.pega.com", datetime.now().strftime('%Y-%m-%d')],
    ["Nintex", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.nintex.com", datetime.now().strftime('%Y-%m-%d')],
    ["Bizagi", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.bizagi.com", datetime.now().strftime('%Y-%m-%d')],
    ["Flowable", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.flowable.com", datetime.now().strftime('%Y-%m-%d')],
    ["Workato", "Workflow/BPM", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.workato.com", datetime.now().strftime('%Y-%m-%d')],
    ["Lakera", "AI Security", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.lakera.com", datetime.now().strftime('%Y-%m-%d')],
    ["Veza", "Other", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.veza.com", datetime.now().strftime('%Y-%m-%d')],
    ["Microsoft Purview", "Other", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD", "Cloud SaaS", "TBD", "TBD", "https://www.microsoftpurview.com", datetime.now().strftime('%Y-%m-%d')],
],

    ["Varonis", "DSPM/DLP", "2005", "New York, NY", "Public (VRNS)", "~$400M", "IPO", "2014", "~$3.5B", "2,000+", 
     "Data Security Platform", "On-prem/Cloud", "Data classification, Access governance, Threat detection", 
     "Comprehensive data security, Legacy platform depth", "https://www.varonis.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Forcepoint", "DLP/SASE", "1994", "Austin, TX", "Private", "N/A", "Acquired by TPG", "2021", "$3B+", "2,500+", 
     "DLP, CASB, ZTNA, SWG", "Hybrid", "Traditional DLP, Web security, ZTNA", 
     "Enterprise DLP leader, SASE integration", "https://www.forcepoint.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Code42", "Insider Risk/DLP", "2001", "Minneapolis, MN", "Private", "$139M", "Series E", "2017", "~$1B", "400+", 
     "Incydr (Insider Risk)", "Cloud SaaS", "Insider threat, Data exfiltration, File activity monitoring", 
     "Insider risk specialization, File-level visibility", "https://www.code42.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Proofpoint", "Email/DLP", "2002", "Sunnyvale, CA", "Private (acquired)", "$469M", "Acquired by Thoma Bravo", "2021", "$12.3B", "4,000+", 
     "Email Security, DLP, CASB", "Cloud SaaS", "Email DLP, People-centric security, Compliance", 
     "Email security leader, People risk intelligence", "https://www.proofpoint.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Digital Guardian", "DLP/EDR", "1996", "Waltham, MA", "Private (acquired)", "N/A", "Acquired by HelpSystems", "2020", "N/A", "500+", 
     "Endpoint DLP, Data Detection & Response", "Endpoint/Cloud", "Endpoint DLP, Managed DLP, DDR platform", 
     "Managed DLP service, Deep endpoint visibility", "https://www.digitalguardian.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Securiti.ai", "DSPM/Privacy", "2019", "San Jose, CA", "Series C", "$267M", "Series C", "2023", "$1B+", "500+", 
     "PrivacyOps, Data Command Center", "Cloud SaaS", "Privacy automation, DSPM, Data governance", 
     "Privacy-first platform, Unified data intelligence", "https://www.securiti.ai", datetime.now().strftime("%Y-%m-%d")],
    
    ["BigID", "DSPM/Privacy", "2016", "New York, NY", "Series E", "$416M", "Series E", "2022", "$1.35B", "400+", 
     "Data Discovery, DSPM, Privacy", "Cloud/Hybrid", "Data discovery, Privacy compliance, Security posture", 
     "ML-driven discovery, Privacy focus", "https://www.bigid.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Normalyze", "DSPM", "2020", "Palo Alto, CA", "Series A", "$26M", "Series A", "2022", "~$100M", "50-100", 
     "Cloud DSPM", "Cloud SaaS", "Cloud data security, Sensitive data discovery, Risk assessment", 
     "Agentless DSPM, Multi-cloud focus", "https://www.normalyze.ai", datetime.now().strftime("%Y-%m-%d")],
    
    ["Wiz", "CNAPP/DSPM", "2020", "Tel Aviv/New York", "Series D", "$1.9B", "Series D", "2024", "$12B", "1,500+", 
     "Cloud Security Platform", "Cloud SaaS", "CNAPP, DSPM, Kubernetes security, Cloud vulnerabilities", 
     "Fastest growing security startup, Comprehensive cloud security", "https://www.wiz.io", datetime.now().strftime("%Y-%m-%d")],
    
    ["Lacework", "CNAPP/DSPM", "2015", "San Jose, CA", "Series F", "$1.8B", "Series F", "2022", "$8.3B", "600+", 
     "Cloud Security Platform", "Cloud SaaS", "CNAPP, DSPM, Cloud workload protection", 
     "Polygraph data platform, Anomaly detection", "https://www.lacework.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Orca Security", "CNAPP/DSPM", "2019", "Tel Aviv/Los Angeles", "Series C", "$550M", "Series C", "2021", "$1.8B", "500+", 
     "Agentless Cloud Security", "Cloud SaaS", "CNAPP, DSPM, Agentless security", 
     "SideScanning tech, Agentless approach", "https://www.orca.security", datetime.now().strftime("%Y-%m-%d")],
    
    ["Cyera", "DSPM", "2021", "Tel Aviv/New York", "Series C", "$300M", "Series C", "2024", "$1.4B", "300+", 
     "DSPM Platform", "Cloud SaaS", "Data security posture, Data discovery, Access intelligence", 
     "Modern DSPM leader, Data-centric approach", "https://www.cyera.io", datetime.now().strftime("%Y-%m-%d")],
    
    ["Laminar", "DSPM", "2020", "Tel Aviv/Boston", "Series A", "$30M", "Series A", "2022", "~$150M", "50-100", 
     "Cloud DSPM", "Cloud SaaS", "Data discovery, Classification, Access control", 
     "Data security graph, Cloud-native", "https://www.laminarsecurity.com", datetime.now().strftime("%Y-%m-%d")],
    
    ["Dig Security", "DSPM", "2021", "Tel Aviv/Boston", "Series A", "$34M", "Series A", "2022", "~$150M", "50-100", 
     "Cloud DSPM", "Cloud SaaS", "Data security, Access governance, Data flow mapping", 
     "Data lineage tracking, Policy automation", "https://www.dig.security", datetime.now().strftime("%Y-%m-%d")],
    
    ["Veza", "Identity/Data Access", "2020", "Los Gatos, CA", "Series B", "$111M", "Series B", "2022", "~$500M", "200+",
     "Access Intelligence Platform", "Cloud SaaS", "Identity governance, Data access, Least privilege",
     "Authorization graph, Identity-to-data visibility", "https://www.veza.com", datetime.now().strftime("%Y-%m-%d")],

    ["Sentra", "DSPM", "2021", "Tel Aviv/New York", "Series B", "$85M", "Series B", "2023", "~$400M", "100-150",
     "Data Security Platform", "Cloud SaaS", "DSPM, Data discovery, Access governance, Threat detection",
     "Data lineage mapping, Real-time data security", "https://www.sentra.io", datetime.now().strftime("%Y-%m-%d")],

    ["Netskope", "CASB/DLP/SASE", "2012", "Santa Clara, CA", "Private", "$1B+", "Series G", "2023", "$7.5B", "2,000+",
     "Security Cloud Platform", "Cloud SaaS", "CASB, DLP, ZTNA, Cloud security",
     "Market leader in CASB, Strong DLP capabilities", "https://www.netskope.com", datetime.now().strftime("%Y-%m-%d")],

    ["Zscaler", "SASE/DLP", "2007", "San Jose, CA", "Public (ZS)", "$1B+", "IPO", "2018", "$28B", "7,000+",
     "Zero Trust Exchange", "Cloud SaaS", "SASE, DLP, Cloud security, Zero trust",
     "SASE category leader, Massive scale", "https://www.zscaler.com", datetime.now().strftime("%Y-%m-%d")],

    ["Microsoft Purview", "DLP/Compliance", "2021", "Redmond, WA", "Microsoft", "N/A", "N/A", "N/A", "N/A", "N/A",
     "Data Governance & Compliance", "Cloud SaaS", "DLP, Compliance, Information protection, eDiscovery",
     "Microsoft 365 integration, Enterprise reach", "https://www.microsoft.com/purview", datetime.now().strftime("%Y-%m-%d")],

    ["Lakera", "AI Security", "2021", "Zurich/San Francisco", "Series A", "$20M", "Series A", "2024", "~$100M", "50-100",
     "GenAI Security Platform", "Cloud SaaS", "AI prompt injection defense, LLM security, GenAI guardrails",
     "Pure-play GenAI security, Prompt injection focus", "https://www.lakera.ai", datetime.now().strftime("%Y-%m-%d")],

    ["Polar Security", "DSPM", "2021", "Tel Aviv", "Series A", "$30M", "Series A", "2023", "~$120M", "50-100",
     "Data Security & Compliance", "Cloud SaaS", "DSPM, Shadow data discovery, Data compliance automation",
     "Agentless architecture, Auto-discovery", "https://www.polarsecurity.io", datetime.now().strftime("%Y-%m-%d")],

    ["Symantec DLP", "DLP", "1982", "Tempe, AZ", "Broadcom (acquired)", "N/A", "Acquired", "2019", "N/A", "20,000+ (Broadcom)",
     "Enterprise DLP Suite", "On-prem/Cloud", "Endpoint DLP, Network DLP, Email DLP, Cloud DLP",
     "Legacy market leader, Enterprise installed base", "https://www.broadcom.com/products/cybersecurity/dlp", datetime.now().strftime("%Y-%m-%d")],

    ["Skyhigh Security", "CASB/DLP", "2011", "Santa Clara, CA", "Symphony Technology", "N/A", "Acquired from McAfee", "2022", "N/A", "500+",
     "CASB & DLP Platform", "Cloud SaaS", "CASB, Cloud DLP, Shadow IT discovery",
     "Former McAfee MVISION, Strong CASB heritage", "https://www.skyhighsecurity.com", datetime.now().strftime("%Y-%m-%d")],

    ["CrowdStrike", "EDR/XDR/Cloud", "2011", "Austin, TX", "Public (CRWD)", "$3B+", "IPO", "2019", "$75B", "9,000+",
     "Falcon Platform", "Cloud/Endpoint", "EDR, XDR, Cloud security, Identity protection, Data protection",
     "Market leader in EDR, Massive scale, Platform approach", "https://www.crowdstrike.com", datetime.now().strftime("%Y-%m-%d")],

    ["SentinelOne", "EDR/XDR", "2013", "Mountain View, CA", "Public (S)", "$1.2B", "IPO", "2021", "$7B", "2,000+",
     "Singularity Platform", "Cloud/Endpoint", "EDR, XDR, Endpoint protection, Data control",
     "AI-powered autonomous EDR, Strong growth", "https://www.sentinelone.com", datetime.now().strftime("%Y-%m-%d")],

    ["SquareX", "Browser Security/DLP", "2022", "San Francisco, CA", "Series A", "$26M", "Series A", "2025", "~$100M", "20-50",
     "Browser Security Platform", "Cloud SaaS", "Browser DLP, Web isolation, Zero trust browsing, Malware protection",
     "Browser-native DLP, Zero-install, Modern approach", "https://www.sqrx.com", datetime.now().strftime("%Y-%m-%d")],

    ["Island", "Enterprise Browser", "2020", "Dallas, TX", "Series E", "$730M", "Series E", "2025", "$4.8B", "500+",
     "Island Enterprise Browser", "Cloud SaaS", "Enterprise browser, Native DLP, Zero trust access, VDI replacement, AI governance",
     "Market leader 15% share, Purpose-built browser, Highest satisfaction", "https://www.island.io", datetime.now().strftime("%Y-%m-%d")],

    ["Talon (Palo Alto)", "Enterprise Browser", "2021", "Tel Aviv, Israel", "Acquired", "$143M", "Acquired by PANW", "2023", "$625M", "100+",
     "Prisma Access Browser", "Cloud SaaS", "Enterprise browser, DLP, BYOD support, Session monitoring",
     "Now part of Palo Alto SASE, Zero-trust enforcement", "https://www.paloaltonetworks.com", datetime.now().strftime("%Y-%m-%d")],

    ["LayerX Security", "Browser Security", "2021", "Tel Aviv, Israel", "Series B", "$145M", "Series B", "2025", "~$600M", "75+",
     "LayerX AI Security Platform", "Cloud SaaS", "Browser-agnostic DLP, GenAI governance, Extension control, SaaS security",
     "Agentless, Works with any browser, AI-powered", "https://www.layerxsecurity.com", datetime.now().strftime("%Y-%m-%d")],

    ["Seraphic Security", "Browser Security", "2020", "Herzliya, Israel", "Series A", "$29M", "Series A", "2025", "~$120M", "49",
     "Enterprise Browser Security", "Cloud SaaS", "Browser-agnostic DLP, Electron app protection, Malware defense",
     "Patented tech, CrowdStrike partnership, 300% ARR growth", "https://www.seraphicsecurity.com", datetime.now().strftime("%Y-%m-%d")],

    ["Surf Security", "Browser-SASE", "2021", "London, UK", "Seed", "Undisclosed", "Seed", "2022", "~$50M", "11-50",
     "Surf Enterprise Browser", "Cloud SaaS", "Browser-SASE, End-to-end encryption, DLP, Phishing protection",
     "First Browser-SASE solution, Founded by CISO/CTO duo", "https://www.surf.security", datetime.now().strftime("%Y-%m-%d")],

    ["Menlo Security", "Browser Isolation", "2013", "Palo Alto, CA", "Series D", "$250M", "Series D", "2019", "$800M", "426+",
     "Menlo Security Platform", "Cloud SaaS", "Remote browser isolation, DLP, Cloud security, Content disarm",
     "Pioneer in RBI, Isolation Core architecture", "https://www.menlosecurity.com", datetime.now().strftime("%Y-%m-%d")],

    ["Camunda", "Workflow/BPM", "2008", "Berlin, Germany", "Series B", "$126M", "Series B", "2021-03", "N/A", "570+",
     "Process Orchestration Platform", "Cloud/On-prem", "BPMN workflows, Process automation, Microservices orchestration, Decision automation",
     "Open-source heritage, Developer-first approach, Microservices focus", "https://camunda.com", datetime.now().strftime("%Y-%m-%d")],

    ["ProcessMaker", "Workflow/BPM", "2000", "Durham, NC", "Series A", "$45M", "Series A", "2021-02", "N/A", "125-199",
     "Low-Code BPM Platform", "Cloud SaaS", "Process automation, Intelligent document processing, Workflow management",
     "Low barrier to entry, Document processing focus", "https://www.processmaker.com", datetime.now().strftime("%Y-%m-%d")],

    ["Appian", "Workflow/Low-Code", "1999", "McLean, VA", "Public (APPN)", "$47.5M", "IPO", "2017-05", "$2.09B", "2,000+",
     "Low-Code Automation Platform", "Cloud SaaS", "Low-code development, Process automation, AI-powered workflow, Data fabric",
     "Market leader low-code, AI-powered automation, Enterprise reach", "https://www.appian.com", datetime.now().strftime("%Y-%m-%d")],

    ["Pega Systems", "BPM/CRM", "1983", "Cambridge, MA", "Public (PEGA)", "$525M", "Post-IPO", "2020-02", "$8.1B", "6,000+",
     "Business Process Platform", "Cloud/On-prem", "BPM, CRM, Process automation, Decision management, AI-powered workflow",
     "Enterprise leader, AI-driven decisioning, Deep industry solutions", "https://www.pega.com", datetime.now().strftime("%Y-%m-%d")],

    ["Nintex", "Workflow Automation", "2006", "Bellevue, WA", "Acquired", "$135M", "Acquired by TPG", "2021-10", "$222M", "1,200+",
     "Process Intelligence Platform", "Cloud SaaS", "Process mapping, Workflow automation, Process intelligence, RPA",
     "Strong Microsoft integration, Process discovery focus", "https://www.nintex.com", datetime.now().strftime("%Y-%m-%d")],

    ["Bizagi", "BPM/Low-Code", "1989", "Wooburn Green, UK", "Private Equity", "$48M", "Private Equity", "2017-09", "N/A", "450+",
     "Low-Code Process Automation", "Cloud/On-prem", "BPM, Process automation, Low-code development, Digital transformation",
     "Rapid deployment, Industry templates, Global presence", "https://www.bizagi.com", datetime.now().strftime("%Y-%m-%d")],

    ["Flowable", "BPM/Workflow", "2016", "Bern, Switzerland", "Private Equity", "N/A", "PE (Ardian)", "N/A", "N/A", "200+",
     "Process Automation Platform", "Cloud/On-prem", "BPMN workflows, Case management, DMN decisions, Process orchestration",
     "Open-source based, Compliance focus, Multi-cloud support", "https://www.flowable.com", datetime.now().strftime("%Y-%m-%d")],

    ["Workato", "iPaaS/Automation", "2013", "Palo Alto, CA", "Series E", "$421M", "Series E", "2021-11", "$5.7B", "1,000+",
     "Enterprise Automation Platform", "Cloud SaaS", "Integration (iPaaS), Workflow automation, API management, Enterprise orchestration",
     "Leading iPaaS, Pre-built connectors, Enterprise automation focus", "https://www.workato.com", datetime.now().strftime("%Y-%m-%d")],
]

for row_data in competitors:
    sheet.append(row_data)

# Data formatting
for row in range(2, len(competitors) + 2):
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Column widths
column_widths = [15, 12, 8, 18, 12, 12, 12, 12, 12, 10, 25, 12, 35, 35, 25, 12]
for i, width in enumerate(column_widths, 1):
    sheet.column_dimensions[get_column_letter(i)].width = width

# Freeze top row
sheet.freeze_panes = 'A2'

# Sheet 2: News & Updates Tracker
news_sheet = wb.create_sheet("News & Updates")

news_headers = [
    "Date", "Company", "Category", "Title", "Summary", "Source/Link", 
    "Impact Level", "Action Required", "Notes", "Status"
]

news_sheet.append(news_headers)

for col_num, header in enumerate(news_headers, 1):
    cell = news_sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

sample_news = [
    [datetime.now().strftime("%Y-%m-%d"), "Nightfall AI", "Product Launch", 
     "Nyx Autonomous DLP Analyst Release", 
     "AI-powered autonomous analyst for DLP event triage and investigation",
     "https://www.nightfall.ai/blog/nyx", "High", "Review feature set", 
     "Differentiator in autonomous investigation", "Reviewed"],
    
    [datetime.now().strftime("%Y-%m-%d"), "Cyera", "Funding", 
     "Series C - $300M at $1.4B valuation", 
     "Major funding round led by Coatue, Accel - largest DSPM funding to date",
     "TechCrunch", "High", "Monitor market positioning", 
     "Shows strong DSPM market validation", "Tracking"],
]

for row_data in sample_news:
    news_sheet.append(row_data)

news_column_widths = [12, 15, 15, 30, 40, 30, 12, 20, 30, 12]
for i, width in enumerate(news_column_widths, 1):
    news_sheet.column_dimensions[get_column_letter(i)].width = width

for row in range(2, len(sample_news) + 2):
    for col in range(1, len(news_headers) + 1):
        cell = news_sheet.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

news_sheet.freeze_panes = 'A2'

# Sheet 3: Product Comparison Matrix
product_sheet = wb.create_sheet("Product Comparison")

product_headers = [
    "Feature/Capability", "Nightfall AI", "Varonis", "Forcepoint", "Code42",
    "Cyera", "BigID", "Wiz", "Netskope", "Zscaler", "Sentra", "Lakera", "CrowdStrike", "SentinelOne", "SquareX", "Island", "LayerX", "Seraphic", "Notes"
]

product_sheet.append(product_headers)

for col_num, header in enumerate(product_headers, 1):
    cell = product_sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

features = [
    ["Shadow AI Detection", "✓", "Limited", "✗", "✗", "Limited", "Limited", "✗", "Limited", "✗", "Limited", "✗", "Limited", "Limited", "✓", "✓", "✓", "Limited",
     "Nightfall's key differentiator"],
    ["GenAI Data Leakage Prevention", "✓", "Partial", "Partial", "✗", "Partial", "Partial", "✗", "Partial", "Partial", "Partial", "✓", "Partial", "Partial", "✓", "✓", "✓", "✓",
     "Critical for modern CISO concerns"],
    ["Cloud-Native DLP", "✓", "✓", "Partial", "✗", "✓", "✓", "✓", "✓", "✓", "✓", "N/A", "✓", "✓", "✓", "✓", "✓", "✓",
     ""],
    ["On-Premise DLP", "✗", "✓", "✓", "✗", "✗", "Partial", "✗", "Partial", "Partial", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗",
     ""],
    ["SaaS App Coverage", "100+", "50+", "Limited", "Limited", "80+", "100+", "Cloud only", "100+", "80+", "100+", "Via API", "Via API", "Via API", "All browsers", "All SaaS", "All browsers", "All browsers",
     ""],
    ["Endpoint DLP", "✗", "✓", "✓", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✓", "✗", "✗", "✗", "✗",
     ""],
    ["Browser DLP", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "Limited", "Limited", "✓", "✓", "✓", "✓",
     "Browser security specialization"],
    ["Email DLP", "Via integrations", "✓", "✓", "✗", "✗", "✗", "✗", "✓", "✓", "✗", "✗", "Limited", "Limited", "Browser-based", "✓", "Browser-based", "Browser-based",
     ""],
    ["Data Discovery", "✓", "✓", "✓", "Limited", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "Limited", "Limited", "Limited", "✓", "Limited", "Limited",
     ""],
    ["Data Classification", "✓ (ML)", "✓", "✓", "Limited", "✓", "✓ (ML)", "✓", "✓ (ML)", "✓", "✓ (ML)", "✗", "✓ (ML)", "✓ (AI)", "✓", "✓ (ML)", "✓ (AI)", "✓",
     ""],
    ["Access Governance", "Limited", "✓", "Limited", "✗", "✓", "✓", "✓", "✓", "Limited", "✓", "✗", "Limited", "Limited", "✗", "✓", "Limited", "Limited",
     ""],
    ["Autonomous Investigation", "✓ (Nyx)", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "Partial", "Partial", "✗", "✗", "Partial (AI)", "✗",
     "Nightfall unique"],
    ["Multi-Cloud Support", "✓", "✓", "Partial", "✗", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "N/A", "✓", "✓", "✓",
     ""],
    ["Browser Agnostic", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "Extension", "Browser Replace", "✓", "✓",
     "LayerX/Seraphic advantage"],
    ["Policy Templates", "✓", "✓", "✓", "Limited", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓",
     ""],
    ["API Integration", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓",
     ""],
    ["Compliance Frameworks", "Multiple", "Multiple", "Multiple", "Limited", "Multiple", "Multiple", "Multiple", "Multiple", "Multiple", "Multiple", "Limited", "Multiple", "Multiple", "Limited", "Multiple", "Multiple", "Multiple",
     ""],
    ["Deployment Time", "Days", "Weeks", "Months", "Weeks", "Days-Weeks", "Days-Weeks", "Days", "Days-Weeks", "Days-Weeks", "Days", "Hours-Days", "Days-Weeks", "Days-Weeks", "Minutes", "Days", "Hours", "Hours",
     ""],
    ["Pricing Model", "Per app/user", "Per user/device", "Per device", "Per device", "Custom", "Custom", "Custom", "Per user", "Per user", "Custom", "API calls", "Per endpoint", "Per endpoint", "Per user", "Per user", "Per user", "Custom",
     ""],
]

for row_data in features:
    product_sheet.append(row_data)

product_column_widths = [30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 30]
for i, width in enumerate(product_column_widths, 1):
    product_sheet.column_dimensions[get_column_letter(i)].width = width

for row in range(2, len(features) + 2):
    for col in range(1, len(product_headers) + 1):
        cell = product_sheet.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

product_sheet.freeze_panes = 'B2'

# Sheet 4: Win/Loss Analysis
winloss_sheet = wb.create_sheet("Win-Loss Analysis")

winloss_headers = [
    "Date", "Opportunity", "Account Name", "Deal Size", "Outcome", 
    "Competitor(s)", "Key Decision Factors", "Our Strengths", "Our Weaknesses",
    "Lessons Learned", "Action Items"
]

winloss_sheet.append(winloss_headers)

for col_num, header in enumerate(winloss_headers, 1):
    cell = winloss_sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

winloss_column_widths = [12, 25, 20, 12, 10, 20, 30, 30, 30, 30, 30]
for i, width in enumerate(winloss_column_widths, 1):
    winloss_sheet.column_dimensions[get_column_letter(i)].width = width

winloss_sheet.freeze_panes = 'A2'

# Sheet 5: Battlecard Quick Reference
battlecard_sheet = wb.create_sheet("Battlecards")

battlecard_headers = [
    "Competitor", "Positioning Statement", "Key Strengths", "Key Weaknesses",
    "Our Advantage", "Trap-Setting Questions", "Proof Points"
]

battlecard_sheet.append(battlecard_headers)

for col_num, header in enumerate(battlecard_headers, 1):
    cell = battlecard_sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

battlecards = [
    ["Varonis",
     "Legacy enterprise data security platform",
     "• Mature platform\n• Strong brand recognition\n• Comprehensive features\n• Large install base",
     "• Complex deployment (weeks/months)\n• High TCO\n• Limited Shadow AI detection\n• Not AI-native",
     "We're AI-native with autonomous investigation (Nyx), deploy in days not months, purpose-built for Shadow AI threats",
     "• How long does deployment typically take?\n• What's your approach to Shadow AI data leakage?\n• Do you have autonomous investigation capabilities?",
     "• Nightfall deploys in days vs. Varonis weeks\n• Nyx autonomous analyst (unique)\n• 100+ SaaS connectors out of box"],

    ["Forcepoint",
     "Enterprise DLP legacy leader",
     "• Long market history\n• Enterprise relationships\n• Broad DLP coverage\n• On-prem strength",
     "• Legacy architecture\n• Complex to manage\n• Slow innovation cycle\n• Weak in cloud-native",
     "Cloud-native architecture, modern UI/UX, faster time-to-value, purpose-built for SaaS/GenAI threats",
     "• How do you handle GenAI data leakage?\n• What's your cloud-native DLP approach?\n• How quickly can you deploy?",
     "• 5x faster deployment\n• Native GenAI protection\n• Modern cloud architecture"],

    ["Cyera",
     "Modern DSPM platform",
     "• Well-funded ($1.4B valuation)\n• Modern DSPM approach\n• Strong data discovery\n• Good investor backing",
     "• Broader than DLP (less focused)\n• No autonomous investigation\n• Limited Shadow AI focus\n• Newer to market",
     "We're laser-focused on DLP with AI-powered autonomous investigation, deeper Shadow AI detection",
     "• How do you handle Shadow AI specifically?\n• Do you have autonomous investigation?\n• What's your DLP heritage?",
     "• Nyx autonomous analyst\n• DLP-first vs. DSPM-first\n• Shadow AI specialization"],

    ["Netskope",
     "CASB/SASE platform with DLP",
     "• Market leader in CASB\n• Strong enterprise presence\n• Well-funded ($7.5B valuation)\n• Broad security platform",
     "• DLP is one of many features\n• Complex pricing\n• Not DLP-specialized\n• No Shadow AI focus",
     "Purpose-built DLP with Shadow AI detection, Nyx autonomous investigation, simpler deployment & pricing",
     "• How focused are you on DLP vs. broader CASB?\n• What's your approach to Shadow AI?\n• Do you have autonomous DLP investigation?",
     "• DLP-first vs. CASB-first\n• Shadow AI specialization\n• Nyx autonomous analyst"],

    ["Zscaler",
     "SASE leader with DLP module",
     "• Massive scale (public company)\n• Zero trust architecture\n• Strong brand\n• Enterprise installed base",
     "• DLP is add-on to SASE\n• Expensive for DLP-only buyers\n• Complex full-stack platform\n• Limited Shadow AI coverage",
     "Focused DLP solution, no need to buy full SASE stack, AI-native with Nyx, purpose-built for Shadow AI",
     "• Can we buy DLP standalone?\n• What's the learning curve for deployment?\n• How do you detect Shadow AI usage?",
     "• Standalone DLP vs. full SASE required\n• Days to deploy vs. weeks\n• Shadow AI detection built-in"],

    ["Sentra",
     "Modern DSPM platform",
     "• Well-funded startup\n• Modern architecture\n• Data lineage capabilities\n• Cloud-native approach",
     "• DSPM-first, not DLP-first\n• No autonomous investigation\n• Limited GenAI protection\n• Newer to market (2021)",
     "DLP-specialized with proven track record, Nyx autonomous investigation, deeper Shadow AI & GenAI protection",
     "• How do you handle real-time DLP?\n• Do you have autonomous investigation?\n• What's your GenAI data protection approach?",
     "• DLP heritage since 2016\n• Nyx autonomous analyst\n• GenAI-specific protections"],

    ["CrowdStrike",
     "EDR/XDR platform with data protection",
     "• Market leader in EDR\n• Massive scale ($75B valuation)\n• Platform breadth\n• Strong brand recognition",
     "• Data protection is add-on to EDR\n• Endpoint-focused, not SaaS DLP\n• No Shadow AI specialization\n• Requires full platform buy-in",
     "Purpose-built SaaS DLP, Shadow AI focus, Nyx autonomous investigation, no need for endpoint agents",
     "• Can we buy DLP standalone without EDR?\n• How do you handle Shadow AI in SaaS apps?\n• Do you have autonomous DLP investigation?",
     "• Cloud-native vs. endpoint-first\n• Shadow AI detection built-in\n• Nyx autonomous analyst"],

    ["SentinelOne",
     "Endpoint security with data control",
     "• Strong EDR/XDR platform\n• AI-powered detection\n• Public company ($7B valuation)\n• Growing rapidly",
     "• Endpoint-focused, limited cloud DLP\n• Data control is basic\n• No SaaS app coverage\n• Not DLP-specialized",
     "Full SaaS DLP coverage (100+ apps), browser & cloud focus, purpose-built for modern data protection",
     "• How many SaaS apps do you cover?\n• What's your browser DLP approach?\n• Do you detect Shadow AI usage?",
     "• 100+ SaaS apps vs. endpoint only\n• Browser-first approach\n• Shadow AI detection"],

    ["SquareX",
     "Browser security with DLP",
     "• Innovative browser-native approach\n• Zero-install deployment\n• Modern UI/UX\n• Fast growing startup",
     "• Browser-only coverage\n• Very early stage ($26M total)\n• Limited enterprise features\n• Extension-based approach",
     "Comprehensive DLP across SaaS, email, endpoints with enterprise features, Nyx autonomous investigation",
     "• What about non-browser data exfiltration?\n• How do you handle email DLP?\n• Do you support enterprise compliance needs?",
     "• Multi-channel DLP vs. browser-only\n• Enterprise-grade features\n• Proven at scale"],

    ["Island",
     "Enterprise browser market leader",
     "• $4.8B valuation, market leader\n• 15% market share in 3 years\n• 200% YoY growth\n• Purpose-built browser",
     "• Requires browser replacement\n• Higher complexity vs. extension\n• Newer company (2020)\n• No SaaS-first DLP heritage",
     "SaaS-native DLP specialist, no browser replacement needed, Nyx autonomous investigation, 100+ SaaS integrations",
     "• Do users need to replace their browser?\n• What's the change management effort?\n• How do you handle Shadow AI in SaaS apps specifically?",
     "• No browser replacement needed\n• DLP-first vs. browser-first\n• Shadow AI specialization"],

    ["LayerX Security",
     "Browser-agnostic security platform",
     "• Works with any browser\n• $145M raised\n• AI-powered approach\n• Agentless deployment",
     "• Browser security focus, not DLP-specialized\n• Limited autonomous investigation\n• Newer to DLP (2021)\n• GenAI focus, less comprehensive DLP",
     "DLP-specialized with proven track record, Nyx autonomous investigation, comprehensive coverage beyond browser",
     "• How do you handle endpoint DLP?\n• What's your autonomous investigation capability?\n• Do you cover email and collaboration tools?",
     "• Comprehensive DLP vs. browser-focused\n• Nyx autonomous analyst\n• Multi-channel coverage"],

    ["Seraphic Security",
     "Browser-agnostic security with DLP",
     "• Browser-agnostic approach\n• CrowdStrike partnership\n• 300% ARR growth\n• Patented technology",
     "• Browser-centric, limited SaaS DLP\n• No autonomous investigation\n• Smaller scale (49 employees)\n• Early stage ($29M raised)",
     "Purpose-built SaaS DLP, Nyx autonomous investigation, enterprise scale, comprehensive feature set",
     "• How many SaaS apps do you natively support?\n• Do you have autonomous DLP investigation?\n• What's your enterprise customer base?",
     "• 100+ SaaS integrations vs. browser focus\n• Nyx autonomous analyst\n• Enterprise proven at scale"],
]

for row_data in battlecards:
    battlecard_sheet.append(row_data)

battlecard_column_widths = [15, 30, 35, 35, 40, 40, 40]
for i, width in enumerate(battlecard_column_widths, 1):
    battlecard_sheet.column_dimensions[get_column_letter(i)].width = width

for row in range(2, len(battlecards) + 2):
    for col in range(1, len(battlecard_headers) + 1):
        cell = battlecard_sheet.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

battlecard_sheet.freeze_panes = 'A2'

wb.save('competitive_intelligence_tracker.xlsx')
print("Competitive Intelligence Tracker created successfully!")
print("File saved: competitive_intelligence_tracker.xlsx")
print(f"Total competitors tracked: {len(competitors)}")
